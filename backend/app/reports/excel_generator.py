"""
DPDP GUI Compliance Scanner - Excel Report Generator

Generates detailed Excel compliance reports using openpyxl.
"""
import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side,
    NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList

from app.models.scan import Scan
from app.models.finding import Finding
from app.models.application import Application


class ExcelReportGenerator:
    """
    Generates detailed Excel compliance reports.

    Worksheets:
    1. Summary - Overview and compliance score
    2. Findings - All findings in tabular format
    3. By Section - Findings grouped by DPDP section
    4. By Severity - Findings grouped by severity
    5. Remediation - Remediation action items
    6. Evidence - Evidence file references
    7. Technical - Technical scan details
    """

    # Color definitions
    COLORS = {
        "header": "0D47A1",  # Navy blue
        "critical": "DC3545",  # Red
        "high": "FD7E14",     # Orange
        "medium": "FFC107",   # Yellow
        "low": "28A745",      # Green
        "light_gray": "F5F5F5",
        "white": "FFFFFF",
    }

    def __init__(
        self,
        scan: Scan,
        application: Application,
        findings: List[Finding],
        include_raw_data: bool = True,
    ):
        self.scan = scan
        self.application = application
        self.findings = findings
        self.include_raw_data = include_raw_data
        self.wb = Workbook()
        self._setup_styles()

    def _setup_styles(self):
        """Setup named styles for the workbook."""
        # Header style
        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(bold=True, color="FFFFFF", size=11)
        self.header_style.fill = PatternFill("solid", fgColor=self.COLORS["header"])
        self.header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.header_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title style
        self.title_style = NamedStyle(name="title")
        self.title_style.font = Font(bold=True, size=16, color=self.COLORS["header"])
        self.title_style.alignment = Alignment(horizontal="left")

        # Subtitle style
        self.subtitle_style = NamedStyle(name="subtitle")
        self.subtitle_style.font = Font(bold=True, size=12)
        self.subtitle_style.alignment = Alignment(horizontal="left")

    async def generate(self) -> io.BytesIO:
        """
        Generate the Excel report.

        Returns:
            BytesIO buffer containing the Excel file
        """
        # Use data passed in during initialization
        scan = self.scan
        findings = self.findings
        application = self.application

        # Remove default sheet
        del self.wb["Sheet"]

        # Create worksheets
        self._create_summary_sheet(scan, application, findings)
        self._create_findings_sheet(findings)
        self._create_by_section_sheet(findings)
        self._create_by_severity_sheet(findings)
        self._create_remediation_sheet(findings)
        self._create_technical_sheet(scan)

        # Save to buffer
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)

        return buffer

    def _create_summary_sheet(
        self,
        scan: Scan,
        application: Application,
        findings: List[Finding],
    ):
        """Create summary worksheet."""
        ws = self.wb.create_sheet("Summary", 0)

        # Title
        ws["A1"] = "DPDP Compliance Scan Report"
        ws["A1"].font = Font(bold=True, size=20, color=self.COLORS["header"])
        ws.merge_cells("A1:F1")

        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(italic=True, color="666666")

        # Application Info
        ws["A4"] = "Application Information"
        ws["A4"].font = Font(bold=True, size=14)

        info_data = [
            ["Application Name", application.name],
            ["URL/Path", application.url or "N/A"],
            ["Type", application.type.value if hasattr(application.type, 'value') else str(application.type)],
            ["Scan Date", (scan.completed_at or scan.created_at).strftime("%Y-%m-%d %H:%M")],
            ["Pages Scanned", scan.pages_scanned],
            ["Scan ID", str(scan.id)],
        ]

        for i, (label, value) in enumerate(info_data, 5):
            ws[f"A{i}"] = label
            ws[f"A{i}"].font = Font(bold=True)
            ws[f"B{i}"] = str(value)

        # Compliance Score
        ws["A12"] = "Compliance Score"
        ws["A12"].font = Font(bold=True, size=14)

        score = scan.overall_score or 0
        ws["B13"] = f"{score}%"
        ws["B13"].font = Font(bold=True, size=24, color=
            self.COLORS["low"] if score >= 80 else
            self.COLORS["medium"] if score >= 60 else
            self.COLORS["critical"]
        )

        # Severity Summary
        ws["A15"] = "Findings by Severity"
        ws["A15"].font = Font(bold=True, size=14)

        severity_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        for finding in findings:
            sev = finding.severity.value if hasattr(finding.severity, 'value') else str(finding.severity)
            if sev in severity_counts:
                severity_counts[sev] += 1

        row = 16
        for severity, count in severity_counts.items():
            ws[f"A{row}"] = severity.capitalize()
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = count
            ws[f"B{row}"].fill = PatternFill("solid", fgColor=self.COLORS.get(severity, "CCCCCC"))
            row += 1

        ws["A21"] = f"Total Findings: {len(findings)}"
        ws["A21"].font = Font(bold=True, size=12)

        # Add pie chart
        if sum(severity_counts.values()) > 0:
            # Data for chart
            ws["E16"] = "Severity"
            ws["F16"] = "Count"
            for i, (sev, count) in enumerate(severity_counts.items(), 17):
                ws[f"E{i}"] = sev.capitalize()
                ws[f"F{i}"] = count

            chart = PieChart()
            chart.title = "Findings by Severity"
            labels = Reference(ws, min_col=5, min_row=17, max_row=20)
            data = Reference(ws, min_col=6, min_row=16, max_row=20)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.width = 10
            chart.height = 8

            ws.add_chart(chart, "D23")

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 10

    def _create_findings_sheet(self, findings: List[Finding]):
        """Create detailed findings worksheet."""
        ws = self.wb.create_sheet("All Findings")

        # Headers
        headers = [
            "ID", "Severity", "DPDP Section", "Check Type",
            "Title", "Description", "Page/Location",
            "Element", "Remediation", "Created At",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=self.COLORS["header"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data rows
        for row, finding in enumerate(findings, 2):
            sev = finding.severity.value if hasattr(finding.severity, 'value') else str(finding.severity)
            check_type = finding.check_type.value if hasattr(finding.check_type, 'value') else str(finding.check_type)

            ws.cell(row=row, column=1, value=str(finding.id)[:8])
            ws.cell(row=row, column=2, value=sev.upper())
            ws.cell(row=row, column=3, value=finding.dpdp_section or "N/A")
            ws.cell(row=row, column=4, value=check_type)
            ws.cell(row=row, column=5, value=finding.title)
            ws.cell(row=row, column=6, value=finding.description)
            ws.cell(row=row, column=7, value=finding.location or "N/A")
            ws.cell(row=row, column=8, value=finding.element_selector or "N/A")
            ws.cell(row=row, column=9, value=finding.remediation or "N/A")
            ws.cell(row=row, column=10, value=finding.created_at.strftime("%Y-%m-%d %H:%M"))

            # Color code severity
            severity_cell = ws.cell(row=row, column=2)
            severity_cell.fill = PatternFill("solid", fgColor=self.COLORS.get(sev, "CCCCCC"))
            if sev in ["critical", "high"]:
                severity_cell.font = Font(bold=True, color="FFFFFF")

            # Alternate row colors
            if row % 2 == 0:
                for col in range(1, 11):
                    if col != 2:  # Skip severity column
                        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=self.COLORS["light_gray"])

        # Set column widths
        widths = [10, 12, 15, 25, 40, 50, 35, 25, 50, 18]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add filter
        ws.auto_filter.ref = f"A1:J{len(findings) + 1}"

    def _create_by_section_sheet(self, findings: List[Finding]):
        """Create findings grouped by DPDP section."""
        ws = self.wb.create_sheet("By Section")

        # Group findings
        sections = {}
        for finding in findings:
            section = finding.dpdp_section or "Other"
            if section not in sections:
                sections[section] = []
            sections[section].append(finding)

        row = 1
        for section, section_findings in sorted(sections.items()):
            # Section header
            ws.cell(row=row, column=1, value=section)
            ws.cell(row=row, column=1).font = Font(bold=True, size=14, color=self.COLORS["header"])
            ws.merge_cells(f"A{row}:E{row}")
            row += 1

            # Column headers
            headers = ["Severity", "Title", "Description", "Page", "Remediation"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=self.COLORS["header"])
            row += 1

            # Findings
            for finding in section_findings:
                sev = finding.severity.value if hasattr(finding.severity, 'value') else str(finding.severity)
                ws.cell(row=row, column=1, value=sev.upper())
                ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=self.COLORS.get(sev, "CCCCCC"))
                ws.cell(row=row, column=2, value=finding.title)
                ws.cell(row=row, column=3, value=finding.description[:200] if finding.description else "")
                ws.cell(row=row, column=4, value=finding.location or "")
                ws.cell(row=row, column=5, value=finding.remediation or "")
                row += 1

            row += 1  # Empty row between sections

        # Set column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 50

    def _create_by_severity_sheet(self, findings: List[Finding]):
        """Create findings grouped by severity."""
        ws = self.wb.create_sheet("By Severity")

        # Group findings
        severities = {"critical": [], "high": [], "medium": [], "low": []}
        for finding in findings:
            sev = finding.severity.value if hasattr(finding.severity, 'value') else str(finding.severity)
            if sev in severities:
                severities[sev].append(finding)

        row = 1
        for severity in ["critical", "high", "medium", "low"]:
            sev_findings = severities[severity]
            if not sev_findings:
                continue

            # Severity header
            ws.cell(row=row, column=1, value=f"{severity.upper()} ({len(sev_findings)} findings)")
            ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=self.COLORS[severity])
            if severity in ["critical", "high"]:
                ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
            ws.merge_cells(f"A{row}:D{row}")
            row += 1

            # Column headers
            headers = ["Section", "Title", "Page", "Remediation"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor=self.COLORS["light_gray"])
            row += 1

            # Findings
            for finding in sev_findings:
                ws.cell(row=row, column=1, value=finding.dpdp_section or "N/A")
                ws.cell(row=row, column=2, value=finding.title)
                ws.cell(row=row, column=3, value=finding.location or "N/A")
                ws.cell(row=row, column=4, value=finding.remediation or "N/A")
                row += 1

            row += 1

        # Set column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 50

    def _create_remediation_sheet(self, findings: List[Finding]):
        """Create remediation action items worksheet."""
        ws = self.wb.create_sheet("Remediation")

        # Title
        ws["A1"] = "Remediation Action Items"
        ws["A1"].font = Font(bold=True, size=16, color=self.COLORS["header"])
        ws.merge_cells("A1:E1")

        # Headers
        headers = ["Priority", "Finding", "DPDP Section", "Remediation Action", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=self.COLORS["header"])

        # Sort findings by severity for priority order
        priority_order = {"critical": 1, "high": 2, "medium": 3, "low": 4}
        sorted_findings = sorted(
            findings,
            key=lambda f: priority_order.get(
                f.severity.value if hasattr(f.severity, 'value') else str(f.severity), 5
            )
        )

        row = 4
        for finding in sorted_findings:
            if not finding.remediation:
                continue

            sev = finding.severity.value if hasattr(finding.severity, 'value') else str(finding.severity)
            priority = sev.upper()

            ws.cell(row=row, column=1, value=priority)
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=self.COLORS.get(sev, "CCCCCC"))
            if sev in ["critical", "high"]:
                ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")

            ws.cell(row=row, column=2, value=finding.title)
            ws.cell(row=row, column=3, value=finding.dpdp_section or "N/A")
            ws.cell(row=row, column=4, value=finding.remediation)
            ws.cell(row=row, column=5, value="Open")  # Status column for tracking

            row += 1

        # Set column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 60
        ws.column_dimensions["E"].width = 12

        # Add data validation for status
        from openpyxl.worksheet.datavalidation import DataValidation
        status_validation = DataValidation(
            type="list",
            formula1='"Open,In Progress,Completed,Deferred"',
            allow_blank=True,
        )
        ws.add_data_validation(status_validation)
        status_validation.add(f"E4:E{row}")

        # Freeze header
        ws.freeze_panes = "A4"

    def _create_technical_sheet(self, scan: Scan):
        """Create technical details worksheet."""
        ws = self.wb.create_sheet("Technical Details")

        # Title
        ws["A1"] = "Technical Scan Details"
        ws["A1"].font = Font(bold=True, size=16, color=self.COLORS["header"])

        # Scan details
        details = [
            ["Scan ID", str(scan.id)],
            ["Scan Type", scan.scan_type.value if hasattr(scan.scan_type, 'value') else str(scan.scan_type)],
            ["Status", scan.status.value if hasattr(scan.status, 'value') else str(scan.status)],
            ["Started At", scan.started_at.strftime("%Y-%m-%d %H:%M:%S") if scan.started_at else "N/A"],
            ["Completed At", scan.completed_at.strftime("%Y-%m-%d %H:%M:%S") if scan.completed_at else "N/A"],
            ["Duration", self._calculate_duration(scan)],
            ["Pages Scanned", scan.pages_scanned],
            ["Findings Count", scan.findings_count],
            ["Compliance Score", f"{scan.overall_score}%" if scan.overall_score else "N/A"],
        ]

        for row, (label, value) in enumerate(details, 3):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))

        # Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 50

    def _calculate_duration(self, scan: Scan) -> str:
        """Calculate scan duration."""
        if scan.started_at and scan.completed_at:
            duration = scan.completed_at - scan.started_at
            minutes = int(duration.total_seconds() / 60)
            seconds = int(duration.total_seconds() % 60)
            return f"{minutes}m {seconds}s"
        return "N/A"
